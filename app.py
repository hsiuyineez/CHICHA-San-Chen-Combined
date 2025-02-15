from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from openpyxl.reader.excel import load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from filelock import FileLock
import shelve
import logging
from discount import *
from datetime import datetime, timedelta
import random
import string
from flask import Flask, render_template, request, redirect, url_for,flash,session,send_from_directory
from Forms import leaveForm, mcForm
from leave import leave
from mc import mc
from Voucher import Voucher
from datetime import datetime, timedelta
from flask import Response
import shelve
import os
from werkzeug.utils import secure_filename

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your_secret_key'

DB_FILE = 'users.db'
filename = "user_data.xlsx"
lockfile = f"{filename}.lock"

@app.route("/")
def home():
    return render_template("base.html")

@app.route("/login")
def customer_login():
    return render_template("login.html")

@app.route("/staff_login")
def staff_login():
    return render_template("staff_login.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]
        phone = request.form['phone']
        birthdate = request.form['birth']
        channels = request.form.getlist("channels")


        hashed_password = generate_password_hash(password)

        with shelve.open(DB_FILE, writeback=True) as db:
            if email in db:
                flash("Email already registered. Please log in.", "error")
                return redirect(url_for("login"))

            db[email] = {
                "name": name,
                "email": email,
                "password": hashed_password,
                'birthdate':birthdate,
                "points": 0,
                "vouchers":[],
                "channels":channels,
                'phone':phone
            }

            #Add 90% off voucher for new user
            voucher = Voucher(
                code="WELCOME90",
                discount=90,
                description="90% off for new customers",
                expiry_date=(datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")  # 7-day expiry
            )
            db[email]["vouchers"].append(voucher.__dict__)  # Store as a dict

        with FileLock(lockfile):
            workbook = load_workbook(filename)
            sheet = workbook.active
            sheet.append([name, email, phone, hashed_password,birthdate, ",".join(channels)])
            workbook.save(filename)

        flash("Registration successful! Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


def is_weekend(date=None):
    if date is None:
        date = datetime.now()
    return date.weekday() in [5, 6]  # Saturday (5) or Sunday (6)

def remove_expired_vouchers(user):
    current_time = datetime.now()
    user['vouchers'] = [
        v for v in user['vouchers']
        if 'expiry_date' not in v or
        (isinstance(v["expiry_date"], str) and datetime.strptime(v["expiry_date"], "%Y-%m-%d %H:%M:%S") > current_time) or
        (isinstance(v["expiry_date"], datetime) and v["expiry_date"] > current_time)
    ]



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        with shelve.open(DB_FILE) as db:
            user = db.get(email)

            if user:
                if "vouchers" not in user:
                    user["vouchers"] = []
                else:
                    remove_expired_vouchers(user)

                    weekend_voucher_code = "WEEKEND50"
                    user['vouchers'] = [v for v in user['vouchers'] if
                                        not (v["code"] == weekend_voucher_code and not is_weekend())]

                db[email] = user

        if user and check_password_hash(user['password'], password):
            session["user"] = user

            if is_weekend():
                voucher = Voucher(
                    code="WEEKEND50",
                    discount=50,
                    description="50% off during weekends",
                )
                if voucher.__dict__ not in user["vouchers"]:
                    with shelve.open(DB_FILE, writeback=True) as db:
                        db[email]["vouchers"].append(voucher.__dict__)
                        session["user"] = db[email]  # Update session with new data


            flash(f"Welcome, {user['name']}!", 'success')
            return redirect(url_for("dashboard"))

        flash("Invalid email or password.","error")

    return render_template("login.html")



@app.route("/dashboard")
def dashboard():
    user = session.get("user")  # Retrieve user info from the session
    if not user:
        flash("Please log in to access the dashboard.", 'error')
        return redirect(url_for("login"))

    # Pass the user object from the session to the template
    return render_template("dashboard.html", user=user)

@app.route("/account_information")
def account_information():
    user = session.get("user")
    if not user:
        flash("Please log in to access account information.", 'error')
        return redirect(url_for("login"))

    user_email = user['email']  # Use user email as identifier

    # Retrieve user's points from shelve
    with shelve.open('user_data') as db:
        user['points'] = db.get(user_email, {}).get('points', 0)

    return render_template("account_information.html", user=user)


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.",'success')
    return redirect(url_for("login"))

@app.route("/about")
def about():
    return render_template("dashboard.html")

def load_workbook(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.append(["Name", "Email", "Phone", "Password", "Birthdate","Channels"])  # Add headers
        workbook.save(filename)
    return workbook

@app.route("/update_details", methods=["GET", "POST"])
def update_details():
    user = session.get("user")
    if not user:
        flash("Please log in to update your details.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        phone = request.form["phone"]

        with shelve.open(DB_FILE, writeback=True) as db:
            db_user = db.get(user["email"])
            if not db_user:
                flash("User not found.", "error")
                return redirect(url_for("account_information"))

            db_user["name"] = name
            db_user["email"] = email
            db_user["phone"] = phone
            db[user["email"]] = db_user

            session["user"] = db_user

        with FileLock(lockfile):
            workbook = load_workbook(filename)
            sheet = workbook.active
            for row in sheet.iter_rows():
                if row[1].value == user["email"]:
                    row[0].value = name
                    row[2].value = phone
                    row[1].value = email
                    break
            workbook.save(filename)

        flash("Details updated successfully.", "success")
        return redirect(url_for("account_information"))

    return render_template("update_details.html", user=user)

@app.route("/loading")
def loading():
    return render_template("loading.html")

@app.route("/redeem_voucher", methods=["POST"])
def redeem_voucher():
    user = session.get("user")
    if not user:
        flash("Please log in to redeem a voucher.", "error")
        return redirect(url_for("login"))

    voucher_code = request.form["voucher_code"]
    purchase_amount = float(request.form["purchase_amount"])

    with shelve.open(DB_FILE, writeback=True) as db:
        db_user = db.get(user["email"])
        if not db_user:
            flash("User not found.", "error")
            return redirect(url_for("dashboard"))

        if "vouchers" not in db_user:
            db_user["vouchers"] = []
            db[user["email"]] = db_user

        voucher = next((v for v in db_user["vouchers"] if v["code"] == voucher_code), None)
        if not voucher:
            flash("Voucher not found or already used.", "error")
            return redirect(url_for("dashboard"))

        discount_amount = (purchase_amount * voucher["discount"]) / 100
        final_amount = purchase_amount - discount_amount

        db_user["vouchers"].remove(voucher)
        db[user["email"]] = db_user
        session["user"] = db_user
        session["final_amount"] = final_amount

    flash(f"Voucher applied! You saved {discount_amount:.2f}. Final amount: {final_amount:.2f}.", "success")
    return redirect(url_for("payment"))

# Menu
@app.route('/menu', methods=['GET', 'POST'])
def menu():
    cart_count = len(session.get('cart', []))

    if 'cart' not in session:
        session['cart'] = []

    if request.method == 'POST':
        location = request.form.get('location', '')
        session['location'] = location

        item_name = request.form.get('name', '')
        item_price = float(request.form.get('price', 0.0))

        # Handle missing item_name gracefully
        if item_name:
            base = request.form.get(f'base_{item_name.lower().replace(" ", "_")}', '')
            sugar = request.form.get(f'sugar_{item_name.lower().replace(" ", "_")}', '')
            ice = request.form.get(f'ice_{item_name.lower().replace(" ", "_")}', '')
            size = request.form.get(f'size_{item_name.lower().replace(" ", "_")}', '')
            remark = request.form.get(f'remark_{item_name.lower().replace(" ", "_")}', '')
            quantity = int(request.form.get(f'quantity_{item_name.lower().replace(" ", "_")}', 1))  # Default quantity to 1
        else:
            # Handle missing item name (e.g., display an error message)
            return render_template('menu.html', error="Item name is missing.")

        item = {
            'name': item_name,
            'price': item_price,
            'base': base,
            'sugar': sugar,
            'ice': ice,
            'size': size,
            'remark': remark,
            'quantity': quantity
        }
        session['cart'].append(item)
        session.modified = True
        return redirect(url_for('menu'))

    return render_template('menu.html', cart_count=cart_count)

@app.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    data = request.json
    item = data.get('item')
    price = data.get('price')
    base = data.get('base')
    sugar = data.get('sugar')
    ice = data.get('ice')
    size = data.get('size')
    remark = data.get('remark')
    quantity = data.get('quantity', 1)

    cart = session.get('cart', [])
    cart.append({
        'name': item,
        'price': float(price),
        'base': base,
        'sugar': sugar,
        'ice': ice,
        'size': size,
        'remark': remark,
        'quantity': int(quantity)
    })
    session['cart'] = cart
    return jsonify(cart_count=len(cart))

@app.route('/cart')
def cart():
    cart = session.get('cart', [])
    location = session.get('location', '')
    subtotal = sum(item['price'] * item['quantity'] for item in cart)
    gst = subtotal * 0.09
    total = subtotal + gst
    return render_template('cart.html', cart=cart, location=location, subtotal=subtotal, gst=gst, total=total)

@app.route('/remove_item/<int:index>', methods=['GET'])
def remove_item(index):
    cart = session.get('cart', [])
    if 0 <= index < len(cart):
        del cart[index]
        session['cart'] = cart
    return redirect(url_for('cart'))

#Payment
@app.route('/payment', methods=['GET', 'POST'])
def payment():
    if request.method == 'POST':
        # Process the payment details here
        cart = session.get('cart', [])
        subtotal = sum(item['price'] * item['quantity'] for item in cart)
        gst = subtotal * 0.09
        total = subtotal + gst

        # Store the total amount as points in shelve
        user = session.get('user')
        if user:
            user_email = user['email']
            with shelve.open('user_data') as db:
                user_data = db.get(user_email, {'points': 0})
                user_data['points'] += total
                db[user_email] = user_data

        # Store final amount in session
        session['final_amount'] = total

    cart = session.get('cart', [])
    subtotal = sum(item['price'] * item['quantity'] for item in cart)
    gst = subtotal * 0.09
    total = subtotal + gst
    final_amount = session.get('final_amount', total)

    return render_template('payment.html', cart=cart, subtotal=subtotal, gst=gst, total=total, final_amount=final_amount)



#Order Summary
@app.route('/order_summary')
def order_summary():
    current_time = datetime.now()
    order_time = session.get('order_time')

    if order_time:
        order_time = datetime.strptime(order_time, '%Y-%m-%d %H:%M')
        if current_time - order_time > timedelta(minutes=10):
            session.pop('order_number', None)
            session.pop('order_time', None)
            session.pop('cart', None)
            return render_template('order_summary.html', order_number=None, location=None, collection_time=None,
                                   cart=[], subtotal=0, gst=0, total=0)

    else:
        order_number = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        session['order_number'] = order_number
        session['order_time'] = current_time.strftime('%Y-%m-%d %H:%M')

    order_number = session.get('order_number')
    location = session.get('location', 'Not selected')
    cart = session.get('cart', [])
    subtotal = sum(item['price'] * item['quantity'] for item in cart)
    gst = subtotal * 0.09
    total = subtotal + gst
    collection_time = (current_time + timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M')

    user = session.get('user')
    if user is None:
        flash('No user ID found in session. Please log in.', 'error')
        return redirect(url_for('login'))

    user_email = user['email']

    # Store the total amount as points in shelve
    with shelve.open('user_data') as db:
        user_data = db.get(user_email, {'points': 0})
        user_data['points'] += total
        db[user_email] = user_data

    return render_template('order_summary.html', cart=cart, order_number=order_number, location=location,
                           collection_time=collection_time, subtotal=subtotal, gst=gst, total=total)



# Contact Us
@app.route('/contact')
def contact():
    return render_template('contact.html')

# Route to handle form submission
@app.route('/submit_contact', methods=['POST'])
def submit_contact():
    # Retrieve form data
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    nature = request.form['nature']
    outlet = request.form['outlet']
    date = request.form['date']
    time = request.form['time']
    message = request.form['message']

    # Store the form data in Shelve
    with shelve.open('contact_data.db', writeback=True) as db:
        if 'contacts' not in db:
            db['contacts'] = []
        db['contacts'].append({
            'name': name,
            'email': email,
            'phone': phone,
            'nature': nature,
            'outlet': outlet,
            'date': date,
            'time': time,
            'message': message,
        })

    flash("Your message has been submitted successfully!", "success")
    return redirect(url_for('success'))

# Route for success page
@app.route('/success')
def success():
    return render_template('success.html')



  # Required for Flask-WTF forms
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}


def get_mc_records(staff_id):
    """ Fetch MC records from shelve database """
    with shelve.open('mc.db', 'r') as db:
        mc_dict = db.get('mc', {})
    return mc_dict.get(staff_id)

mc_records = []

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return "No file part"

        file = request.files['file']

        if file.filename == '':
            return "No selected file"

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            return f"File {filename} uploaded successfully!"

    return render_template('upload.html')


@app.route('/staff_login', methods=['GET', 'POST'])
def stafflogin():
    if request.method == 'POST':
        staff_id = request.form['staff_id']
        password = request.form['password']

        # Open the staff database
        db = shelve.open('staff.db', 'r')  # Open in read mode
        staff_dict = db.get('staff', {})

        if staff_id in staff_dict and staff_dict[staff_id] == password:
            session['staff_id'] = staff_id  # Save staff ID in session
            db.close()
            return redirect(url_for('homepage'))  # Redirect to home page
        else:
            flash('Invalid staff ID or password. Please try again.', 'danger')
            db.close()

    return render_template('staff_login.html')  # Serve the login page

@app.route('/homepage', methods=['GET', 'POST'])
def homepage():
    if 'staff_id' in session:  # Check if staff ID is stored in session
        staff_id = session['staff_id']
        return render_template('homepage.html', staff_id=staff_id)
    else:
        flash('Please log in to access the home page.', 'warning')
        return redirect(url_for('stafflogin'))

@app.route('/stafflogout')
def stafflogout():
    session.pop('staff_id', None)  # Remove staff ID from session
    flash('You have been logged out.', 'info')
    return redirect(url_for('stafflogin'))



@app.route('/leave', methods=['GET', 'POST'])
def leave_view():
    leave_form = leaveForm(request.form)
    if 'staff_id' in session:
        leave_form.staff_id.data = session['staff_id']

    if request.method == 'POST' and leave_form.validate():
        db = shelve.open("leave.db", "c", writeback=True)
        leave_dict = db.get("leave", {})

        # Use staff_id as the key
        staff_id = leave_form.staff_id.data
        leave_dict[staff_id] = leave(
            staff_id,
            leave_form.starting_date.data,
            leave_form.end_date.data,
            leave_form.reason.data
        )

        db["leave"] = leave_dict  # Save back
        db.close()

        flash("Leave submitted successfully!", "success")
        return redirect(url_for('leave_view'))

    return render_template('leave.html', form=leave_form)

@app.route('/mc', methods=['GET', 'POST'])
def mc_view():
    mc_form = mcForm(request.form)
    if 'staff_id' in session:
        mc_form.staff_id.data = session['staff_id']

    if request.method == 'POST' and mc_form.validate():
        db = shelve.open("mc.db", "c", writeback=True)
        mc_dict = db.get("mc", {})

        # Use staff_id as the key
        staff_id = mc_form.staff_id.data

        mc_dict[staff_id] = mc(
            staff_id,
            mc_form.starting_date.data,
            mc_form.end_date.data,
            mc_form.proof.data
        )

        db["mc"] = mc_dict  # Save back
        db.close()

        flash("MC submitted successfully!", "success")
        return redirect(url_for('mc_view'))

    return render_template('mc.html', form=mc_form)
@app.route('/retrieveleave', methods=['GET', 'POST'])
def retrieve_leave():
    staff_id = session.get('staff_id')  # Get the logged-in staff's ID from the session
    if not staff_id:
        flash('Please log in to view your records.', 'danger')
        return redirect(url_for('stafflogin'))  # Redirect if no staff is logged in

    # Open the shelve database in read mode
    db = shelve.open('leave.db', 'r')

    # Get all leave records
    leave_dict = db.get('leave', {})

    # Filter leave records for the logged-in staff member
    staff_leave_records = [record for record in leave_dict.values() if record.get_staff_id() == staff_id]

    # Close the shelve database
    db.close()

    # Return the filtered leave records to the template for rendering
    return render_template('retrieveleave.html', leave_list=staff_leave_records)


@app.route('/retrievemc')
def retrieve_mc():
    staff_id = session.get('staff_id')  # Get the logged-in staff's ID from the session
    if not staff_id:
        flash('Please log in to view your records.', 'danger')
        return redirect(url_for('stafflogin'))  # Redirect if no staff is logged in

    # Open the shelve database in read mode
    db = shelve.open('mc.db', 'r')

    # Get all MC records
    mc_dict = db.get('mc', {})

    # Retrieve the MC record for the logged-in staff (if exists)
    staff_mc_record = mc_dict.get(staff_id)  # Returns None if no record exists

    # Close the shelve database
    db.close()

    if staff_mc_record:
        # Return the MC record to the template for rendering
        return render_template('retrievemc.html', mc_list=[staff_mc_record])
    else:
        flash("No MC record found.", "warning")
        return render_template('retrievemc.html', mc_list=[])

@app.route('/updateleave/<int:id>/', methods=['GET', 'POST'])
def update_leave(id):
    update_leave_form = leaveForm(request.form)

    db = shelve.open('leave.db', 'c')
    leave_dict = db.get('leave', {})
    db.close()
    print(f"Debug: All leave record IDs: {list(leave_dict.keys())}")

    l = leave_dict.get(str(id))  # Normalize the key to string

    if not l:
        return f"Leave record with ID {id} not found.", 404

    if request.method == 'POST' and update_leave_form.validate():
        db = shelve.open('leave.db', 'w')
        leave_dict = db['leave']
        # Update record
        l.set_staff_id(update_leave_form.staff_id.data)
        l.set_starting_date(update_leave_form.starting_date.data)
        l.set_end_date(update_leave_form.end_date.data)
        l.set_reason(update_leave_form.reason.data)
        leave_dict[str(id)] = l  # Ensure key is str
        db['leave'] = leave_dict
        db.close()
        flash("Leave updated successfully!", "success")

        # Redirect after POST
        return redirect(url_for('leave_view'))

    # Pre-populate form
    update_leave_form.staff_id.data = l.get_staff_id()
    update_leave_form.starting_date.data = l.get_starting_date()
    update_leave_form.end_date.data = l.get_end_date()
    update_leave_form.reason.data = l.get_reason()
    return render_template('updateleave.html', form=update_leave_form)


@app.route('/deleteleave/<int:id>', methods=['POST'])
def delete_leave(id):
    db = shelve.open('leave.db', 'w')
    leave_dict = db.get('leave', {})
    leave_dict.pop(str(id), None)  # Avoid KeyError, normalize key to string
    db['leave'] = leave_dict
    db.close()
    return redirect(url_for('retrieve_leave'))


@app.route('/updatemc/<int:id>/', methods=['GET', 'POST'])
def update_mc(id):
    update_mc_form = mcForm(request.form)

    db = shelve.open('mc.db', 'c')
    mc_dict = db.get('mc', {})
    db.close()

    # Debugging output
    print(f"Debug: All keys in mc_dict = {list(mc_dict.keys())}")
    print(f"Debug: Attempting to retrieve record with key = '{str(id)}'")

    m = mc_dict.get(str(id))  # Ensure ID is a string

    if not m:
        return f"MC record with ID {id} not found.", 404

    if request.method == 'POST' and update_mc_form.validate():
        db = shelve.open('mc.db', 'w')
        mc_dict = db['mc']
        # Update record
        m.set_staff_id(update_mc_form.staff_id.data)
        m.set_starting_date(update_mc_form.starting_date.data)
        m.set_end_date(update_mc_form.end_date.data)
        m.set_proof(update_mc_form.proof.data)
        mc_dict[str(id)] = m  # Ensure key is str
        db['mc'] = mc_dict
        db.close()
        flash("MC updated successfully!", "success")

        # Redirect after POST
        return redirect(url_for('mc_view'))

    else:
        db = shelve.open('mc.db', 'r')
        mc_dict = db['mc']
        db.close()
        m = mc_dict.get(str(id))
        if m:
            # Pre-populate form
            update_mc_form.staff_id.data = m.get_staff_id()
            update_mc_form.starting_date.data = m.get_starting_date()
            update_mc_form.end_date.data = m.get_end_date()
            update_mc_form.proof.data = m.get_proof()
        return render_template('updatemc.html', form=update_mc_form)


@app.route('/deletemc/<int:id>', methods=['POST'])
def delete_mc(id):
    key = str(id)  # Ensure ID is a string
    db = shelve.open('mc.db', 'w')
    mc_dict = db.get('mc', {})
    if key in mc_dict:
        mc_dict.pop(key)
        db['mc'] = mc_dict
        db.close()
        return redirect(url_for('retrieve_mc'))
    else:
        db.close()
        return f"MC record with ID {id} is not able to delete as it is removed in the database", 404


@app.route('/normalize_keys')
def normalize_keys():
    db = shelve.open('mc.db', 'w')
    mc_dict = db.get('mc', {})
    updated_mc_dict = {str(k): v for k, v in mc_dict.items()}  # Normalize keys to strings
    db['mc'] = updated_mc_dict
    db.close()
    return "Keys in mc.db have been normalized to strings."


@app.route('/staff_discount')
def staff_discount():
    staff_id = session.get('staff_id')
    if not staff_id:
        return "Please log in first."

    vouchers = Voucher.load_vouchers()
    collected_vouchers = Voucher.load_staff_vouchers(staff_id)

    # Ensure the loaded collected_vouchers has the expected structure
    print("Collected Vouchers Data:", collected_vouchers)  # Debugging: Check the structure of loaded collected_vouchers

    return render_template('staff_discount.html', vouchers=vouchers.values(), collected_vouchers=collected_vouchers)
@app.route('/collect/<code>', methods=['POST'])
def collect_voucher(code):
    staff_id = session.get('staff_id')
    if not staff_id:
        flash('Please log in to collect vouchers.', 'danger')
        return redirect(url_for('staff_discount'))

    vouchers = Voucher.load_vouchers()
    collected_vouchers = Voucher.load_staff_vouchers(staff_id)  # Load only this staff's vouchers

    if code in vouchers:
        current_time = datetime.now()

        # Ensure collected_vouchers[code] is a dictionary
        if code not in collected_vouchers or isinstance(collected_vouchers[code], dict):
            collected_vouchers[code] = {
                'collected_at': current_time,  # Save the datetime under 'collected_at'
                'collected_by': staff_id       # Optionally store the staff_id who collected it
            }
            Voucher.save_staff_vouchers(staff_id, collected_vouchers)  # Save for this staff ONLY
            flash(f'Congratulations! You have collected {vouchers[code].description}.', 'success')
        else:
            next_available = collected_vouchers[code]['collected_at'] + timedelta(days=30)
            flash(f'You have already collected this voucher. Available again on {next_available.strftime("%Y-%m-%d")}.',
                  'warning')
    else:
        flash('Voucher not found!', 'danger')

    return redirect(url_for('staff_discount'))

@app.route('/search', methods=['GET'])
def search():
    query = request.args.get('query', '').lower()  # Get the search query and make it case-insensitive

    if 'leave records' in query:
        return redirect(url_for('retrieve_leave'))  # Redirect to the leave records page
    elif 'mc records' in query:
        return redirect(url_for('retrieve_mc'))  # Redirect to the MC records page
    elif any(keyword in query for keyword in ['staff discount', 'discount', 'vouchers', 'voucher']):
        return redirect(url_for('staff_discount'))  # Redirect to the staff discount page
    elif 'submit mc' in query:
        return redirect(url_for('mc_view'))
    elif 'submit leave' in query:
        return redirect(url_for('leave_view'))  # Redirect to the submit MC/Leave page
    else:
        return redirect(url_for('homepage'))



@app.route('/view_proof/<staff_id>')
def view_proof(staff_id):
    print(f"Viewing proof for staff_id: {staff_id}")  # Debugging
    mc_record = get_mc_records(staff_id)

    if mc_record:
        print(f"Proof data found: {mc_record.get_proof()}")  # Debugging
        proof_data = mc_record.get_proof()

        if proof_data:
            file_name = os.path.basename(proof_data)
            return send_from_directory(app.config['UPLOAD_FOLDER'], file_name)
    else:
        print(f"MC record not found for staff_id: {staff_id}")  # Debugging

    return "Proof not found", 404

if __name__ == '__main__':
    app.run(debug=True)
# 1. generate code and description (10%, 20% off), collection (status, collected or not)
# 2. voucher database

